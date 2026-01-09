from fastapi import FastAPI, UploadFile, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from dotenv import load_dotenv
from openai import OpenAI
from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook
import os
import shutil
import json
import re
import uuid
from typing import List

# -------------------------------------------------
# Init
# -------------------------------------------------

load_dotenv()

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
if not OPENAI_API_KEY:
    raise RuntimeError("OPENAI_API_KEY mangler")

client = OpenAI(api_key=OPENAI_API_KEY)

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = "uploads"
DATABLAD_DIR = "datablad"

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(DATABLAD_DIR, exist_ok=True)

# -------------------------------------------------
# Sesjonskontekst (per chat)
# -------------------------------------------------

SESSION_CONTEXT: dict[str, dict] = {}

# -------------------------------------------------
# Fil-lesing
# -------------------------------------------------

def read_pdf(path: str) -> str:
    reader = PdfReader(path)
    return "\n".join(page.extract_text() or "" for page in reader.pages)

def read_docx(path: str) -> str:
    doc = Document(path)
    return "\n".join(p.text for p in doc.paragraphs)

def read_excel(path: str) -> str:
    wb = load_workbook(path, data_only=True)
    sheet = wb.active
    rows = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i > 50:
            break
        rows.append(" | ".join(str(c) if c else "" for c in row))
    return "\n".join(rows)

# -------------------------------------------------
# Datablad – Indeksering
# -------------------------------------------------

def chunk_text(text: str, size: int = 500, overlap: int = 100):
    chunks = []
    i = 0
    while i < len(text):
        chunks.append(text[i:i + size])
        i += size - overlap
    return chunks

def index_datasheet(pdf_path: str, vendor: str):
    text = read_pdf(pdf_path)
    chunks = chunk_text(text)

    data = {
        "vendor": vendor,
        "source": os.path.basename(pdf_path),
        "chunks": chunks
    }

    with open(pdf_path.replace(".pdf", ".json"), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def index_all_datasheets(vendor: str):
    vendor_dir = os.path.join(DATABLAD_DIR, vendor)
    if not os.path.exists(vendor_dir):
        raise FileNotFoundError(f"Mangler mappe {vendor_dir}")

    for f in os.listdir(vendor_dir):
        if f.lower().endswith(".pdf"):
            index_datasheet(os.path.join(vendor_dir, f), vendor)

# -------------------------------------------------
# Produktidentifikasjon
# -------------------------------------------------

PRODUCT_REGEX = re.compile(r"\b[A-Z]{2,4}\d{3,5}[-\.]?\d*\b")

def extract_product_ids(text: str) -> List[str]:
    return list(set(PRODUCT_REGEX.findall(text.upper())))

def match_products_by_filename(products: List[str], vendor: str) -> List[str]:
    vendor_dir = os.path.join(DATABLAD_DIR, vendor)
    matches = []

    for f in os.listdir(vendor_dir):
        if not f.lower().endswith(".pdf"):
            continue

        fname = f.lower().replace(" ", "")
        for p in products:
            if p.lower().replace(".", "").replace("-", "") in fname:
                matches.append(f)
                break

    return matches

# -------------------------------------------------
# RAG – forbedret søk
# -------------------------------------------------

def tokenize(text: str):
    return re.findall(r"\w+", text.lower())

def score_chunk(query_tokens, chunk_text):
    tokens = tokenize(chunk_text)
    return sum(tokens.count(t) for t in query_tokens)

def search_datasheets(query: str, vendor: str, max_hits: int = 4):
    vendor_dir = os.path.join(DATABLAD_DIR, vendor)
    if not os.path.exists(vendor_dir):
        return [], []

    query_tokens = tokenize(query)
    scored = []

    for f in os.listdir(vendor_dir):
        if not f.endswith(".json"):
            continue

        try:
            with open(os.path.join(vendor_dir, f), "r", encoding="utf-8") as fh:
                data = json.load(fh)

            for chunk in data.get("chunks", []):
                score = score_chunk(query_tokens, chunk)
                if score > 0:
                    scored.append({
                        "score": score,
                        "text": chunk,
                        "source": data["source"]
                    })
        except Exception:
            continue

    scored.sort(key=lambda x: x["score"], reverse=True)

    hits = []
    sources = []
    for item in scored[:max_hits]:
        hits.append(f"[{item['source']}]\n{item['text']}")
        if item["source"] not in sources:
            sources.append(item["source"])

    return hits, sources

def load_chunks_from_sources(vendor: str, sources: List[str], max_chunks: int = 4):
    vendor_dir = os.path.join(DATABLAD_DIR, vendor)
    chunks = []

    for f in os.listdir(vendor_dir):
        if not f.endswith(".json"):
            continue

        with open(os.path.join(vendor_dir, f), "r", encoding="utf-8") as fh:
            data = json.load(fh)

        if data["source"] in sources:
            for chunk in data["chunks"]:
                chunks.append(f"[{data['source']}]\n{chunk}")
                if len(chunks) >= max_chunks:
                    return chunks

    return chunks

# -------------------------------------------------
# Health
# -------------------------------------------------

@app.get("/")
def root():
    return {"status": "API running"}

# -------------------------------------------------
# Indekser Siemens
# -------------------------------------------------

@app.post("/index/siemens")
def index_siemens():
    index_all_datasheets("Siemens")
    return {"status": "Indeksering fullført", "vendor": "Siemens"}

# -------------------------------------------------
# Chat /ask
# -------------------------------------------------

@app.post("/ask")
async def ask(request: Request):
    session_id = request.headers.get("X-Session-ID") or str(uuid.uuid4())
    SESSION_CONTEXT.setdefault(session_id, {"sources": []})

    content_type = request.headers.get("content-type", "")
    question = None
    role = "Ukjent"
    file = None

    if "application/json" in content_type:
        body = await request.json()
        question = body.get("question")
        role = body.get("role", role)
    elif "multipart/form-data" in content_type:
        form = await request.form()
        question = form.get("question")
        role = form.get("role", role)
        file = form.get("file")
    else:
        raise HTTPException(status_code=415, detail="Unsupported content type")

    if not question:
        raise HTTPException(status_code=400, detail="Question is required")

    # -------- 1. Finn produkt-ID-er --------
    products = extract_product_ids(question)

    # -------- 2. RAG-søk --------
    rag_chunks, rag_sources = search_datasheets(question, "Siemens")

    # -------- 3. Filnavn-fallback --------
    if products:
        filename_hits = match_products_by_filename(products, "Siemens")
        for f in filename_hits:
            if f not in rag_sources:
                rag_sources.append(f)

    # -------- 4. Gjenbruk sesjonskontekst --------
    if not rag_sources and len(question.split()) < 6:
        rag_sources = SESSION_CONTEXT[session_id].get("sources", [])

    # -------- 5. Last chunks --------
    if rag_sources:
        rag_chunks = load_chunks_from_sources("Siemens", rag_sources)
        SESSION_CONTEXT[session_id]["sources"] = rag_sources

    rag_context = "\n\n".join(rag_chunks)[:3000]

    # -------- 6. Vedlegg --------
    attachment_text = ""
    if file and file.filename:
        path = os.path.join(UPLOAD_DIR, file.filename)
        with open(path, "wb") as f:
            shutil.copyfileobj(file.file, f)

        ext = file.filename.lower().split(".")[-1]
        if ext == "pdf":
            attachment_text = read_pdf(path)
        elif ext in ["docx", "doc"]:
            attachment_text = read_docx(path)
        elif ext in ["xlsx", "xls"]:
            attachment_text = read_excel(path)

    # -------- 7. Prompt --------
    messages = [
        {
            "role": "system",
            "content": (
                "Du er en faglig KI-assistent for prosjektstøtte innen byggautomasjon.\n"
                f"Brukerrolle: {role}\n"
                "Svar presist, teknisk korrekt og basert på dokumentasjon."
            )
        }
    ]

    if rag_context:
        messages.append({
            "role": "system",
            "content": f"Relevante utdrag fra datablad:\n{rag_context}"
        })

    if attachment_text:
        messages.append({
            "role": "system",
            "content": f"Innhold fra vedlegg:\n{attachment_text[:2000]}"
        })

    messages.append({"role": "user", "content": question})

    # -------- 8. OpenAI --------
    resp = client.chat.completions.create(
        model="gpt-4.1-mini",
        messages=messages,
        timeout=30
    )

    return {
        "answer": resp.choices[0].message.content,
        "sources": rag_sources,
        "session_id": session_id,
        "used_rag": bool(rag_sources)
    }
